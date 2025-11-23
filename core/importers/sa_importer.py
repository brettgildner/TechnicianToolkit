from __future__ import annotations
import openpyxl
from datetime import datetime
from core.logic import compute_duration
from core.models.service_activity_model import ServiceActivity
import re
import unicodedata

"""
Controls the 'Import from Excel' function for the Service Activity page.
"""

def import_from_excel(self):
    # Imports service activity from Excel file; normalizes header names (fixes whitespace,
    # newlines, weird Unicode), accepts multiple header variations (aliases/typo fixes),
    # skips blank/meaningless rows safely and computes call duration if missing.
    from PySide6.QtWidgets import QFileDialog, QMessageBox, QInputDialog

    # Select file
    path, _ = QFileDialog.getOpenFileName(
        self, "Select Service Activity Excel File", "", "Excel Files (*.xlsx)"
    )
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path)
        sheet_names = wb.sheetnames

        # Choose sheet
        sheet_name, ok = QInputDialog.getItem(
            self,
            "Select Sheet",
            "Choose which sheet to import:",
            sheet_names,
            0,
            False,
        )
        if not ok:
            wb.close()
            return

        ws = wb[sheet_name]

        # Header normalization function
        def normalize_header(h: str) -> str:
            if not h:
                return ""
            h = unicodedata.normalize("NFKC", str(h))
            h = re.sub(r"\s+", " ", h)  # collapse all whitespace
            return h.strip().lower()

        # Read and normalize headers
        raw_header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))

        print("\nDEBUG – Raw Excel headers with repr():")
        for c in raw_header_cells:
            print(repr(c.value))

        excel_headers = [normalize_header(c.value) for c in raw_header_cells]

        # Header aliases & typo fixes
        HEADER_ALIASES = {
            "departure time": ["departure time", "depature time", "depature time", "depature  time"],
            "call duration": ["call duration", "call  duration"],
            "technician last name": ["technician last name", "technician", "tech last name"],
            "service log / comments": ["service log / comments", "comments", "service log"],
            "serial number": ["serial number", "serial #", "serial\nnumber"],
            "arrival time": ["arrival time", "arrival\ntime"],
        }

        # Reverse alias mapping > excel header to canonical key
        alias_lookup = {}
        for canonical, variants in HEADER_ALIASES.items():
            for v in variants:
                alias_lookup[normalize_header(v)] = canonical

        # Build internal > DB field map
        header_map = {
            "area": "area",
            "customer": "customer",
            "serial number": "serial_number",
            "meter": "meter",
            "description of malfunction": "malfunction",
            "arrival date": "arrival_date",
            "arrival time": "arrival_time",
            "remedial action performed": "remedial_action",
            "qty": "quantity",
            "part replaced": "part_replaced",
            "departure date": "departure_date",
            "departure time": "departure_time",
            "call duration": "call_duration",
            "technician last name": "technician",
            "service log / comments": "comments",
        }

        # Normalize Excel headers to canonical names (via alias lookup)
        normalized_excel_headers = []
        for h in excel_headers:
            normalized_excel_headers.append(alias_lookup.get(h, h))

        # Required columns
        required_headers = list(header_map.keys())
        missing = [h for h in required_headers if h not in normalized_excel_headers]

        if missing:
            QMessageBox.critical(
                self,
                "Import Error",
                "The file is missing expected columns:\n" + ", ".join(missing)
            )
            wb.close()
            return

        # Row validation helper
        PRIMARY_FIELDS = [
            "area",
            "customer",
            "malfunction",
            "arrival_date",
            "departure_date",
        ]

        def row_has_real_data(row_dict):
            # Return 'True' if any important field has meaningful content.
            for key in PRIMARY_FIELDS:
                v = row_dict.get(key)
                if v not in (None, "", " "):
                    return True
            return False

        # Import data rows
        added = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip Excel noise rows
            if not any(cell not in (None, "", " ") for cell in row):
                continue
            excel_row = dict(zip(normalized_excel_headers, row))

            # Skip rows without meaningful service activity data
            if not row_has_real_data(excel_row):
                continue

            # Normalize dates/times
            def norm(v):
                if isinstance(v, datetime):
                    # Pure date > convert to YYYY-MM-DD
                    if v.time() == datetime.min.time():
                        return v.date().isoformat()
                    return str(v)  # datetime with time
                return v

            clean = {}
            for excel_name, internal_name in header_map.items():
                clean[internal_name] = norm(excel_row.get(excel_name))

            # Inject user
            clean["user"] = self.user

            # Compute missing duration
            clean["call_duration"] = compute_duration(
                clean.get("arrival_date"),
                clean.get("arrival_time"),
                clean.get("departure_date"),
                clean.get("departure_time"),
            )

            # Insert record
            ServiceActivity.create(**clean)
            added += 1

        wb.close()
        self.load_items()

        # Notify dashboard that service activity changed
        from ui.signals.model_signals import model_signals
        model_signals.service_activity_changed.emit()

        QMessageBox.information(
            self,
            "Import Complete",
            f"Successfully imported {added} service activities from '{sheet_name}'."
        )

    except Exception as e:
        QMessageBox.critical(self, "Import Error", f"An error occurred:\n{e}")