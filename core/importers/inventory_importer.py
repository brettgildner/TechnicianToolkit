from __future__ import annotations
import openpyxl
from datetime import datetime
from typing import List, Tuple
from core.models.inventory_model import InventoryItem

"""
Controls the 'Import from Excel' function for the Inventory page.
"""

REQUIRED_HEADERS = [
    "part number",
    "qty",
    "part location",
    "model",
    "part description",
    "quarterly inventory verification date",
    "days since last verification",
    "notes",
    "today's date",
]

def import_inventory_from_excel(path: str, sheet_name: str, user) -> Tuple[int, List[str]]:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    # Read header row
    headers = [
        str(c.value).strip().lower() if c.value else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    # Check required headers
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        wb.close()
        return 0, missing  # UI will handle the error

    added = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue  # skip blank rows

        row_dict = dict(zip(headers, row))

        raw_date = row_dict.get("quarterly inventory verification date")
        clean_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date

        item_data = {
            "part_number": row_dict.get("part number"),
            "quantity": row_dict.get("qty"),
            "part_location": row_dict.get("part location"),
            "model": row_dict.get("model"),
            "part_description": row_dict.get("part description"),
            "quarterly_inventory_verification_date": clean_date,
            "notes": row_dict.get("notes"),
            "category_id": row_dict.get("category") if "category" in headers else None,
            "user": user,
        }

        # Skip if the row is empty other than user
        if not any(v for k, v in item_data.items() if k != "user"):
            continue

        InventoryItem.create(**item_data)
        added += 1

    wb.close()
    return added, []  # empty missing list -> success
