from __future__ import annotations
from datetime import date, datetime
from typing import List, Optional
import os
from PySide6.QtCore import Qt, QModelIndex, QTimer, QDate
from PySide6.QtGui import QCursor
from PySide6.QtWidgets import (QWidget,QVBoxLayout,QHBoxLayout,QLabel,QPushButton,QMessageBox,QDialog,
                               QLineEdit,QDateEdit,QFileDialog,QHeaderView,QGroupBox,QGridLayout)
import openpyxl
from core.logic import get_current_user
from core.models.expense_model import ExpenseReportInfo, ExpenseEntry as DBExpenseEntry
from ui.components.base_tables.base_table_model import BaseTableModel
from ui.components.base_tables.base_table_view import BaseTableView
from ui.components.action_buttons.action_buttons_delegate import ActionButtonsDelegate
from ui.components.dialogs.expense_help_dialog import ExpenseHelpDialog
from ui.components.dialogs.expense_header_dialog import ExpenseHeaderDialog
from ui.forms.expense_entry_form import ExpenseEntryForm
from core.models.expense_report_header_model import ExpenseReportHeader
from core.models.expense_line_model import ExpenseLine

"""
This page implements the full Expense Report Page, a complete workflow for creating, editing, deleting, 
managing, and exporting employee expense reports. It loads and persists data through SQLite models, 
provides a metadata header section, a detailed expenses table with action buttons, dialogs for editing 
header info or individual expense lines, and functionality to export the entire report into a preformatted 
Excel template. It also ensures consistent UI styling, manages date ranges, column sizing, and keeps the 
model in sync with both the database and on-screen widgets. Overall, this page is the user’s central 
interface for monthly expense reporting.

ui.pages.expense_report_page.py index:

class ExpenseTableModel(): Simple expense table data model.
class ExpenseReportPage(): Main UI for managing expense reports.
- def __init__(): Build full expense page UI.
- def _qdate_to_date(): Convert QDate > python date.
- def _sync_header_from_widgets(): Copy UI metadata > header dataclass.
- def _apply_header_to_widgets(): Copy header dataclass > UI fields.
    - def set_qdate(): Set QDateEdit safely.
- def _actions_logical_index(): Find index of actions column.
- def _refresh_model(): Refresh table model & widths.
- def _ensure_actions_width(): Force action column proper size.
- def edit_header(): Open header edit dialog.
- def _load_header_from_db(): Load saved header metadata.
    - def set_date(): Apply DB date > widget.
- def _save_header_to_db(): Persist header metadata.
- def add_expense(): Open new expense entry form.
- def _on_edit_clicked(): Edit existing expense line.
- def _on_delete_clicked(): Delete an expense line.
- def _on_double_click(): Double-click row > edit.
- def _load_expenses_from_db(): Fetch all expenses for user.
- def _entry_to_dataclass(): Convert DB row > ExpenseLine.
- def export_to_excel(): Fill Excel template with report.
    - def ensure_date(): Convert header dates reliably.
- def clear_all_entries(): Delete all expenses for user.
- def show_help_dialog(): Show help instructions.
"""

# Table model: Expenses
class ExpenseTableModel(BaseTableModel):
    # Simple wrapper around BaseTableModel for DBExpenseEntry objects.
    pass

# Main Expense Report Page
class ExpenseReportPage(QWidget):
    # Matches the look of MileagePage, uses a table and action buttons, persists data via SQLite (DBExpenseEntry
    # & ExpenseReportInfo), and exports to a preformatted Excel template (.xlsx ONLY, NOT .xls)
    def __init__(self, parent=None, controller=None):
        super().__init__(parent)
        self.controller = controller
        self.user = get_current_user() or "default_user"
        self.expenses: List[DBExpenseEntry] = []
        self.header_info: Optional[ExpenseReportInfo] = None
        self.header = ExpenseReportHeader()

        # Layout
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 10, 20, 10)
        root.setSpacing(10)

        header_label = QLabel("Expense Report")
        header_label.setStyleSheet("color: #ffffff; font: bold 28px 'Segoe UI';")
        root.addWidget(header_label, alignment=Qt.AlignLeft)

        # Control Bar
        controls = QHBoxLayout()
        controls.setSpacing(10)

        # Edit header (gray)
        btn_header = QPushButton("Edit Report Header Info")
        btn_header.setFixedWidth(180)
        btn_header.clicked.connect(self.edit_header)
        btn_header.setCursor(QCursor(Qt.PointingHandCursor))
        btn_header.setToolTip(
            "This only needs to be completed before exporting to your blank "
            "'Expense Report' Excel file at the end of the month."
        )
        btn_header.setStyleSheet("""
            QPushButton {
                background-color: #CCAA00;
                color: #ffffff;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #E6C300;
            }
        """)

        # Add expense (green)
        btn_add = QPushButton("+ Add Expense")
        btn_add.setFixedWidth(160)
        btn_add.clicked.connect(self.add_expense)
        btn_add.setCursor(QCursor(Qt.PointingHandCursor))
        btn_add.setStyleSheet("""
            QPushButton {
                background-color: #00cc55;
                color: #ffffff;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #00aa44;
            }
        """)

        # Export (dark blue)
        btn_export = QPushButton("Export to Excel")
        btn_export.setFixedWidth(160)
        btn_export.clicked.connect(self.export_to_excel)
        btn_export.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #004c99;
                color: #ffffff;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #003d7a;
            }
        """)

        # Clear (red)
        btn_clear = QPushButton("Clear All Entries")
        btn_clear.setFixedWidth(160)
        btn_clear.clicked.connect(self.clear_all_entries)
        btn_clear.setCursor(QCursor(Qt.PointingHandCursor))
        btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #8b0000;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #a00000;
            }
        """)

        # Help dialog (orange)
        btn_help = QPushButton("Help")
        btn_help.setFixedWidth(120)
        btn_help.clicked.connect(self.show_help_dialog)
        btn_help.setCursor(QCursor(Qt.PointingHandCursor))
        btn_help.setStyleSheet("""
            QPushButton {
                background-color: #cc6600;
                color: #ffffff;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #995000;
            }
        """)

        # Add to toolbar
        controls.addWidget(btn_add)
        controls.addWidget(btn_export)
        controls.addWidget(btn_header)
        controls.addWidget(btn_clear)
        controls.addWidget(btn_help)
        controls.addStretch()

        root.addLayout(controls)

        # Metadata box (always visible)
        metadata_box = QGroupBox("Expense Report Info")
        metadata_box.setStyleSheet("""
            QGroupBox {
                font: bold 16px 'Segoe UI';
                color: #00ff99;
                border: 2px solid #00ff99;
                border-radius: 8px;
                margin-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                top: -4px;
                padding: 0 4px;
            }
            QLabel {
                color: #ffffff;
            }
        """)
        meta_layout = QGridLayout(metadata_box)
        meta_layout.setVerticalSpacing(8)
        meta_layout.setHorizontalSpacing(12)

        self.name_edit = QLineEdit()
        self.employee_edit = QLineEdit()
        self.phone_edit = QLineEdit()
        self.mail_team_edit = QLineEdit()
        self.group_edit = QLineEdit()
        self.division_edit = QLineEdit()
        self.dest_edit = QLineEdit()

        self.report_date_edit = QDateEdit()
        self.report_date_edit.setCalendarPopup(True)

        self.begin_date_edit = QDateEdit()
        self.begin_date_edit.setCalendarPopup(True)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)

        self.bc_edit = QLineEdit()
        self.account_edit = QLineEdit()

        # Autofill begin/end dates to current month
        today_q = QDate.currentDate()
        self.report_date_edit.setDate(today_q)
        self.begin_date_edit.setDate(QDate(today_q.year(), today_q.month(), 1))
        self.end_date_edit.setDate(QDate(today_q.year(), today_q.month(), today_q.daysInMonth()))

        # Layout header fields
        meta_layout.addWidget(QLabel("Name"), 0, 0)
        meta_layout.addWidget(self.name_edit, 0, 1)
        meta_layout.addWidget(QLabel("Employee #"), 0, 2)
        meta_layout.addWidget(self.employee_edit, 0, 3)

        meta_layout.addWidget(QLabel("Telephone #"), 1, 0)
        meta_layout.addWidget(self.phone_edit, 1, 1)
        meta_layout.addWidget(QLabel("Mail/Team"), 1, 2)
        meta_layout.addWidget(self.mail_team_edit, 1, 3)

        meta_layout.addWidget(QLabel("Group"), 2, 0)
        meta_layout.addWidget(self.group_edit, 2, 1)
        meta_layout.addWidget(QLabel("Division"), 2, 2)
        meta_layout.addWidget(self.division_edit, 2, 3)

        meta_layout.addWidget(QLabel("Destination / Purpose"), 3, 0)
        meta_layout.addWidget(self.dest_edit, 3, 1, 1, 3)

        meta_layout.addWidget(QLabel("Report Date"), 4, 0)
        meta_layout.addWidget(self.report_date_edit, 4, 1)
        meta_layout.addWidget(QLabel("Begin"), 4, 2)
        meta_layout.addWidget(self.begin_date_edit, 4, 3)

        meta_layout.addWidget(QLabel("End"), 5, 0)
        meta_layout.addWidget(self.end_date_edit, 5, 1)
        meta_layout.addWidget(QLabel("B/C"), 5, 2)
        meta_layout.addWidget(self.bc_edit, 5, 3)

        meta_layout.addWidget(QLabel("Account #"), 6, 0)
        meta_layout.addWidget(self.account_edit, 6, 1)

        root.addWidget(metadata_box)

        # Sync initial header from widgets
        self._sync_header_from_widgets()

        # Load from DB (if previously saved)
        self._load_header_from_db()

        # Table
        self.columns = [
            "expense_date",
            "destination",
            "miles",
            "rental",
            "air_cash",
            "air",
            "hotel",
            "meals",
            "ent_bus_mtgs",
            "parking",
            "telephone",
            "misc",
            "explanation",
            "actions",
        ]
        self.column_labels = {
            "expense_date": "Date",
            "destination": "Destination",
            "miles": "# of miles",
            "rental": "Rental",
            "air_cash": "Air-Cash",
            "air": "Air",
            "hotel": "Hotel",
            "meals": "Meals",
            "ent_bus_mtgs": "Ent. Bus. Mtgs.",
            "parking": "Parking",
            "telephone": "Telephone",
            "misc": "Misc.",
            "explanation": "Explanation",
            "actions": "Actions",
        }

        self.table = BaseTableView(self)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(False)
        self.table.setStyleSheet("""
            QTableView {
                background-color: #1a1a1a;
                color: #e0e0e0;
                gridline-color: #333333;
                selection-background-color: #00cc88;
                alternate-background-color: #222222;
            }
            QHeaderView::section {
                background-color: #2a2a2a;
                color: #ffffff;
                font-weight: bold;
                padding: 6px;
                border: none;
            }
        """)
        root.addWidget(self.table, stretch=1)

        # Model
        self.base_model = ExpenseTableModel([], self.columns, self.column_labels)
        self.table.setModel(self.base_model)

        # Load expenses from DB now that model exists
        self._load_expenses_from_db()

        # Actions column delegate (Edit/Delete)
        actions_col = self.columns.index("actions")
        self.actions_delegate = ActionButtonsDelegate(
            parent=self.table,
            button_keys=["edit", "delete"],
            on_edit=self._on_edit_clicked,
            on_delete=self._on_delete_clicked,
        )
        self.table.setItemDelegateForColumn(actions_col, self.actions_delegate)
        self._ensure_actions_width()

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(False)

        # Ensure actions column has enough width
        QTimer.singleShot(0, self._ensure_actions_width)
        self.base_model.modelReset.connect(self._ensure_actions_width)
        self.base_model.layoutChanged.connect(self._ensure_actions_width)

        # Double-click row > edit
        self.table.doubleClicked.connect(self._on_double_click)

    # Helpers for header syncing
    def _qdate_to_date(self, qd: QDate) -> Optional[date]:
        if not qd.isValid():
            return None
        return date(qd.year(), qd.month(), qd.day())

    def _sync_header_from_widgets(self):
        # Update self.header from the always-visible widgets.
        self.header.name = self.name_edit.text().strip()
        self.header.employee_number = self.employee_edit.text().strip()
        self.header.telephone_number = self.phone_edit.text().strip()
        self.header.mail_team = self.mail_team_edit.text().strip()
        self.header.group = self.group_edit.text().strip()
        self.header.division = self.division_edit.text().strip()
        self.header.destination_purpose = self.dest_edit.text().strip()
        self.header.bc = self.bc_edit.text().strip()
        self.header.account_number = self.account_edit.text().strip()

        self.header.report_date = self._qdate_to_date(self.report_date_edit.date())
        self.header.start_date = self._qdate_to_date(self.begin_date_edit.date())
        self.header.end_date = self._qdate_to_date(self.end_date_edit.date())

    def _apply_header_to_widgets(self):
        # Push self.header data back into the visible widgets.
        self.name_edit.setText(self.header.name)
        self.employee_edit.setText(self.header.employee_number)
        self.phone_edit.setText(self.header.telephone_number)
        self.mail_team_edit.setText(self.header.mail_team)
        self.group_edit.setText(self.header.group)
        self.division_edit.setText(self.header.division)
        self.dest_edit.setText(self.header.destination_purpose)
        self.bc_edit.setText(self.header.bc)
        self.account_edit.setText(self.header.account_number)

        def set_qdate(edit: QDateEdit, d: Optional[date]):
            if d is None:
                edit.setDate(QDate.currentDate())
            else:
                edit.setDate(QDate(d.year, d.month, d.day))

        set_qdate(self.report_date_edit, self.header.report_date)
        set_qdate(self.begin_date_edit, self.header.start_date)
        set_qdate(self.end_date_edit, self.header.end_date)

    # Helper: refresh table from self.expenses
    def _actions_logical_index(self, width: int = 200) -> int:
        # Find index of 'actions' column in all_columns list.
        try:
            return self.columns.index("actions")
        except ValueError:
            return -1

    def _refresh_model(self):
        self.base_model.set_items(self.expenses)
        self.table.resizeColumnsToContents()
        QTimer.singleShot(0, self._ensure_actions_width)

    def _ensure_actions_width(self):
        # Ensures the actions column width matches the delegate's requirements.
        li = self._actions_logical_index()
        if li < 0:
            return

        header = self.table.horizontalHeader()

        # Determine correct width from delegate
        delegate = self.table.itemDelegateForColumn(li)
        if hasattr(delegate, "required_width"):
            width = delegate.required_width()
        else:
            width = 220

        header.setSectionResizeMode(li, QHeaderView.ResizeMode.Interactive)

        # Enforce proper width
        if header.sectionSize(li) < width:
            header.resizeSection(li, width)

    # Header dialog and DB persistence
    def edit_header(self):
        # Open the popup dialog, then sync back to widgets and DB.
        # Make sure header reflects latest widget values
        self._sync_header_from_widgets()

        dlg = ExpenseHeaderDialog(self, self.header)
        if dlg.exec() == QDialog.Accepted:
            dlg.apply_to_header()
            self._apply_header_to_widgets()
            self._save_header_to_db()

    def _load_header_from_db(self):
        info = ExpenseReportInfo.get_for_user(self.user)
        self.header_info = info

        # If nothing stored, sync widgets with header dataclass and return
        if info is None:
            self._sync_header_from_widgets()
            return

        # Populate the UI widgets
        self.name_edit.setText(info.name or "")
        self.employee_edit.setText(info.employee_number or "")
        self.phone_edit.setText(info.telephone_number or "")
        self.mail_team_edit.setText(info.mail_team or "")
        self.group_edit.setText(info.group_name or "")
        self.division_edit.setText(info.division or "")
        self.dest_edit.setText(info.destination_purpose or "")

        def set_date(widget, value):
            if isinstance(value, date):
                widget.setDate(QDate(value.year, value.month, value.day))
            else:
                widget.setDate(QDate.currentDate())

        set_date(self.report_date_edit, info.report_date)
        set_date(self.begin_date_edit, info.start_date)
        set_date(self.end_date_edit, info.end_date)

        self.bc_edit.setText(info.bc or "")
        self.account_edit.setText(info.account_number or "")

        # Sync the header dataclass using the converted DB values
        self.header.name = info.name or ""
        self.header.employee_number = info.employee_number or ""
        self.header.telephone_number = info.telephone_number or ""
        self.header.mail_team = info.mail_team or ""
        self.header.group = info.group_name or ""
        self.header.division = info.division or ""
        self.header.destination_purpose = info.destination_purpose or ""
        self.header.bc = info.bc or ""
        self.header.account_number = info.account_number or ""

        self.header.report_date = info.report_date
        self.header.start_date = info.start_date
        self.header.end_date = info.end_date

    def _save_header_to_db(self):
        # Save current metadata widgets to DB and sync into self.header.
        self._sync_header_from_widgets()

        info = ExpenseReportInfo.upsert_for_user(
            user=self.user,
            name=self.header.name,
            employee_number=self.header.employee_number,
            telephone_number=self.header.telephone_number,
            mail_team=self.header.mail_team,
            group_name=self.header.group,
            division=self.header.division,
            destination_purpose=self.header.destination_purpose,
            report_date=self.header.report_date,
            start_date=self.header.start_date,
            end_date=self.header.end_date,
            bc=self.header.bc,
            account_number=self.header.account_number,
        )
        self.header_info = info

    # Add / Edit / Delete entries
    def add_expense(self):
        dlg = ExpenseEntryForm(self, entry=None)
        if dlg.exec() == QDialog.Accepted:
            new_entry = dlg.to_entry()

            # Persist to DB
            created = DBExpenseEntry.create(
                user=self.user,
                expense_date=new_entry.expense_date,
                destination=new_entry.destination,
                miles=new_entry.miles,
                rental=new_entry.rental,
                air_cash=new_entry.air_cash,
                air=new_entry.air,
                hotel=new_entry.hotel,
                meals=new_entry.meals,
                ent_bus_mtgs=new_entry.ent_bus_mtgs,
                parking=new_entry.parking,
                telephone=new_entry.telephone,
                misc=new_entry.misc,
                explanation=new_entry.explanation,
            )
            # Refresh full list from DB to keep IDs in sync
            self._load_expenses_from_db()

    def _on_edit_clicked(self, source_row: int):
        if source_row < 0 or source_row >= len(self.expenses):
            return

        existing = self.expenses[source_row]
        # Convert DB model to UI dataclass
        existing_line = self._entry_to_dataclass(existing)

        dlg = ExpenseEntryForm(self, entry=existing_line)
        if dlg.exec() == QDialog.Accepted:
            updated = dlg.to_entry()

            # Update DB
            DBExpenseEntry.update(
                entry_id=existing.id,
                user=self.user,
                expense_date=updated.expense_date,
                destination=updated.destination,
                miles=updated.miles,
                rental=updated.rental,
                air_cash=updated.air_cash,
                air=updated.air,
                hotel=updated.hotel,
                meals=updated.meals,
                ent_bus_mtgs=updated.ent_bus_mtgs,
                parking=updated.parking,
                telephone=updated.telephone,
                misc=updated.misc,
                explanation=updated.explanation,
            )
            # Refresh from DB
            self._load_expenses_from_db()

    def _on_delete_clicked(self, source_row: int):
        # Delete the selected expense entry.
        if source_row < 0 or source_row >= len(self.expenses):
            return

        entry = self.expenses[source_row]

        confirm = QMessageBox.question(
            self,
            "Delete Expense",
            "Are you sure you want to delete this expense entry?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        DBExpenseEntry.delete(entry.id, user=self.user)
        self._load_expenses_from_db()

    def _on_double_click(self, index: QModelIndex):
        if not index.isValid():
            return
        row = index.row()
        self._on_edit_clicked(row)

    def _load_expenses_from_db(self):
        self.expenses = DBExpenseEntry.get_all_for_user(self.user)
        self.base_model.set_items(self.expenses)
        self.table.resizeColumnsToContents()
        QTimer.singleShot(0, self._ensure_actions_width)

    def _entry_to_dataclass(self, model_obj: DBExpenseEntry) -> ExpenseLine:
        return ExpenseLine(
            expense_date=model_obj.expense_date,
            destination=model_obj.destination,
            miles=model_obj.miles,
            rental=model_obj.rental,
            air_cash=model_obj.air_cash,
            air=model_obj.air,
            hotel=model_obj.hotel,
            meals=model_obj.meals,
            ent_bus_mtgs=model_obj.ent_bus_mtgs,
            parking=model_obj.parking,
            telephone=model_obj.telephone,
            misc=model_obj.misc,
            explanation=model_obj.explanation,
        )

    # Export to Excel (XLSX)
    def export_to_excel(self):
        if not self.expenses:
            QMessageBox.information(self, "No Data", "There are no expense entries to export.")
            return

        # Ask where to save the filled-out report
        out_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Expense Report",
            "expense_report.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not out_path:
            return

        try:
            # Path to the preformatted template
            template_path = os.path.join("assets", "Forms", "Expense_Report.xlsx")

            if not os.path.exists(template_path):
                QMessageBox.critical(
                    self,
                    "Template Missing",
                    f"Could not find template file:\n{template_path}\n\n"
                    "Please update template_path in ExpenseReportPage."
                )
                return

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # Save latest metadata from widgets to DB
            self._save_header_to_db()

            info = self.header_info or ExpenseReportInfo.get_for_user(self.user)
            if info is None:
                QMessageBox.warning(self, "Missing Info", "Please fill out the report info before exporting.")
                return

            # Convert DB date strings to python dates
            start = ExpenseReportInfo._to_date(info.start_date)
            end = ExpenseReportInfo._to_date(info.end_date)

            if not start or not end:
                QMessageBox.warning(self, "Missing Dates", "Start and End dates are required for export.")
                return

            # Filter items within date range
            export_items = []
            for e in self.expenses:
                d = DBExpenseEntry._to_date(e.expense_date)
                if d and start <= d <= end:
                    export_items.append(e)

            if not export_items:
                QMessageBox.information(self, "No Items", "No expenses fall within the selected date range.")
                return

            # Sync widgets to header dataclass
            self._sync_header_from_widgets()
            h = self.header

            # Convert header dates if needed
            def ensure_date(value):
                if isinstance(value, date):
                    return value
                if isinstance(value, str) and value.strip():
                    try:
                        return datetime.strptime(value, "%Y-%m-%d").date()
                    except ValueError:
                        return None
                return None

            h.report_date = ensure_date(h.report_date)
            h.start_date = ensure_date(h.start_date)
            h.end_date = ensure_date(h.end_date)

            # Row 5 — LABELS (always static)
            ws["A5"] = "Destination/Purpose"
            ws["E5"] = "Date"
            ws["F5"] = "Begin"
            ws["G5"] = "End"
            ws["H5"] = "B/C"
            ws["I5"] = "Account Number"

            # Row 6 — VALUES (user input)
            ws["A6"] = h.destination_purpose or ""
            ws["E6"] = h.report_date.strftime("%m/%d/%Y") if h.report_date else ""
            ws["F6"] = h.start_date.strftime("%m/%d/%Y") if h.start_date else ""
            ws["G6"] = h.end_date.strftime("%m/%d/%Y") if h.end_date else ""
            ws["H6"] = h.bc or ""
            ws["I6"] = h.account_number or ""

            # Additional top-of-sheet info
            ws["A2"] = h.name or "Employee Name"
            ws["C2"] = h.employee_number or "Employee Number"
            ws["F2"] = h.telephone_number or "Telephone Number"
            ws["H2"] = h.mail_team or "Mail/Team"
            ws["I2"] = h.group or "Group"
            ws["J2"] = h.division or "Division"

            # Write detail rows (expenses) starting at row 9
            start_row = 9
            for idx, entry in enumerate(export_items):
                r = start_row + idx
                d = DBExpenseEntry._to_date(entry.expense_date)

                ws[f"A{r}"] = d.strftime("%m/%d/%Y") if d else ""
                ws[f"B{r}"] = entry.destination
                ws[f"C{r}"] = entry.miles or 0
                ws[f"D{r}"] = entry.rental or 0
                ws[f"E{r}"] = entry.air_cash or 0
                ws[f"F{r}"] = entry.air or 0
                ws[f"G{r}"] = entry.hotel or 0
                ws[f"H{r}"] = entry.meals or 0
                ws[f"I{r}"] = entry.ent_bus_mtgs or 0
                ws[f"J{r}"] = entry.parking or 0
                ws[f"K{r}"] = entry.telephone or 0
                ws[f"L{r}"] = entry.misc or 0
                ws[f"M{r}"] = entry.explanation

            # Save file
            wb.save(out_path)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Expense report exported to:\n{out_path}",
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"An error occurred while exporting:\n{e}"
            )

    # Clear all entries (for next month)
    def clear_all_entries(self):
        if not self.expenses:
            return
        confirm = QMessageBox.question(
            self,
            "Clear All Entries",
            "This will remove all expense entries for this user.\n\n"
            "Make sure you exported first. Continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return
        DBExpenseEntry.delete_all_for_user(self.user)
        self._load_expenses_from_db()

    # Create the Help button & help dialog
    def show_help_dialog(self):
        dlg = ExpenseHelpDialog(self)
        dlg.exec()