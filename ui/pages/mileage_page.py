from __future__ import annotations
import os
import openpyxl
from typing import List, Optional
from datetime import datetime, date
from PySide6.QtWidgets import (QWidget,QVBoxLayout,QHBoxLayout,QLabel,QPushButton,QMessageBox,QFileDialog,
                               QDialog,QFormLayout,QComboBox,QLineEdit,QDialogButtonBox,QHeaderView,QGroupBox,
                               QGridLayout,QDateEdit)
from PySide6.QtCore import Qt, QModelIndex, QTimer, QDate
from PySide6.QtGui import QCursor
from core.logic import get_current_user, clear_all_mileage_entries
from core.models.mileage_model import MileageEntry
from ui.components.base_tables.base_table_model import BaseTableModel
from ui.components.dialogs.mileage_help_dialog import MileageHelpDialog
from ui.components.filters.filter_proxy_models import MileageFilterProxy
from ui.components.base_tables.base_table_view import BaseTableView
from ui.components.action_buttons.action_buttons_delegate import ActionButtonsDelegate
from ui.forms.mileage_form import MileageForm

"""
This page defines the user interface for the Mileage Tracker, providing tools to view, add, edit, filter, 
delete, and export mileage entries for the current user. It builds a full form for employee, vehicle, and 
date-range information, displays mileage records in a sortable table with edit/delete action buttons, and 
computes miles driven dynamically through a custom table model. Users can filter entries, clear all 
entries, and export selected date-range data into a preformatted Excel mileage-expense template. The page 
manages all interactions between the UI, database models, and export logic while maintaining a polished, 
styled interface.

ui.pages.mileage_page.py index:

class MileageTableModel(): Table model for mileage that adds a computed miles_driven column.
 - def get_value(): Compute the miles driven for display
class MileagePage(): Displays the mileage tracker interface with entry management and export tools.
 - def __init__(): Build UI and load entries.
 - def load_items(): Load items into table
 - def refresh_table(): Reload table data.
 - def _map_proxy_index_to_item(): Helper: map rows/indexes to db items
 - def _map_source_row_to_item(): Helper: map source rows to db items
 - def _on_edit_clicked(): Edit selected entry.
 - def _on_delete_clicked(): Delete selected entry.
 - def open_add_form(): Open new entry form.
 - def edit_item(): Open edit form.
 - def delete_item(): Delete entry with confirmation.
 - def _on_double_click(): Double-click to edit.
 - def open_filter_dialog(): Open filtering dialog.
     - def on_reset(): Reset filters
 - def export_to_excel(): Export report to Excel template.
     - def to_date(): Collect entries in date range.
 - def clear_all_entries(): Clear all entries
 - def _actions_logical_index(): Find actions column.
 - def _ensure_actions_width(): Resize actions column.
 - def show_help_dialog(): Open help dialog.
"""

# Table model for mileage that adds a computed miles_driven column.
class MileageTableModel(BaseTableModel):
    def get_value(self, item: MileageEntry, column_key: str):
        # Compute the miles driven for display
        if column_key == "miles_driven":
            try:
                # Convert values to floats and subtract
                return float(item.end_miles) - float(item.start_miles)
            except Exception:
                return ""  # Show blank if improper numbers

        # For all other columns, use the normal method
        return super().get_value(item, column_key)

class MileagePage(QWidget):
# Displays the mileage tracker interface with entry management and export tools.
    def __init__(self, parent=None, controller=None):
        super().__init__(parent)
        self.controller = controller
        self.user = get_current_user() or "default_user"

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 10, 20, 10)
        root.setSpacing(10)

        header = QLabel("Mileage Tracker")
        header.setStyleSheet("color: #ffffff; font: bold 28px 'Segoe UI';")
        root.addWidget(header, alignment=Qt.AlignLeft)

        controls = QHBoxLayout()
        controls.setSpacing(10)

        # Add (green)
        btn_add = QPushButton("+ Add Mileage Entry")
        btn_add.setFixedWidth(200)
        btn_add.clicked.connect(self.open_add_form)
        btn_add.setCursor(QCursor(Qt.PointingHandCursor))
        btn_add.setStyleSheet("""
            QPushButton {
                background-color: #00cc55;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #00aa44;
            }
        """)

        # Filters (gray)
        btn_filters = QPushButton("Filters")
        btn_filters.setFixedWidth(120)
        btn_filters.clicked.connect(self.open_filter_dialog)
        btn_filters.setCursor(QCursor(Qt.PointingHandCursor))
        btn_filters.setStyleSheet("""
            QPushButton {
                background-color: #444444;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #555555;
            }
        """)

        # Export (blue)
        btn_export = QPushButton("Export to Excel")
        btn_export.setFixedWidth(160)
        btn_export.clicked.connect(self.export_to_excel)
        btn_export.setCursor(QCursor(Qt.PointingHandCursor))
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #004c99;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #003d7a;
            }
        """)

        # Clear all (red)
        btn_clear = QPushButton("Clear All Entries")
        btn_clear.setFixedWidth(150)
        btn_clear.clicked.connect(self.clear_all_entries)
        btn_clear.setCursor(QCursor(Qt.PointingHandCursor))
        btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #8b0000;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #a00000;
            }
        """)

        # Help (orange)
        btn_help = QPushButton("Help")
        btn_help.setFixedWidth(100)
        btn_help.clicked.connect(self.show_help_dialog)
        btn_help.setCursor(QCursor(Qt.PointingHandCursor))
        btn_help.setStyleSheet("""
            QPushButton {
                background-color: #cc6600;
                color: white;
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:hover {
                background-color: #995000;
            }
        """)

        # Add to layout
        controls.addWidget(btn_add)
        controls.addWidget(btn_export)
        controls.addWidget(btn_filters)
        controls.addWidget(btn_clear)
        controls.addWidget(btn_help)
        controls.addStretch()

        root.addLayout(controls)


        # Header info box
        header_box = QGroupBox("Mileage Report Info")
        header_box.setStyleSheet("""
            QGroupBox {
                font: bold 16px 'Segoe UI';
                color: #00ff99;
                border: 2px solid #00ff99;
                border-radius: 8px;
                margin-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                top: -4px;
                padding: 0 4px;
            }
            QLabel {
                color: #ffffff;
            }
        """)
        grid = QGridLayout(header_box)
        grid.setVerticalSpacing(8)
        grid.setHorizontalSpacing(12)

        # Employee / contact info
        self.m_emp_name_edit = QLineEdit()
        self.m_emp_number_edit = QLineEdit()
        self.m_phone_edit = QLineEdit()
        self.m_mail_team_edit = QLineEdit()

        # Vehicle info
        self.m_vehicle_make_edit = QLineEdit()
        self.m_vehicle_model_edit = QLineEdit()
        self.m_plate_edit = QLineEdit()

        # Date range (begin / end)
        self.m_begin_date_edit = QDateEdit()
        self.m_begin_date_edit.setCalendarPopup(True)
        self.m_end_date_edit = QDateEdit()
        self.m_end_date_edit.setCalendarPopup(True)

        # Default begin/end = current month
        today_q = QDate.currentDate()
        begin_q = QDate(today_q.year(), today_q.month(), 1)
        end_q = QDate(today_q.year(), today_q.month(), today_q.daysInMonth())
        self.m_begin_date_edit.setDate(begin_q)
        self.m_end_date_edit.setDate(end_q)

        # Layout:
        # Row 0: Employee Name | Employee #
        grid.addWidget(QLabel("Employee Name"), 0, 0)
        grid.addWidget(self.m_emp_name_edit, 0, 1)
        grid.addWidget(QLabel("Employee #"), 0, 2)
        grid.addWidget(self.m_emp_number_edit, 0, 3)

        # Row 1: Telephone | Mail/Team
        grid.addWidget(QLabel("Telephone #"), 1, 0)
        grid.addWidget(self.m_phone_edit, 1, 1)
        grid.addWidget(QLabel("Mail/Team"), 1, 2)
        grid.addWidget(self.m_mail_team_edit, 1, 3)

        # Row 2: Vehicle Make | Model
        grid.addWidget(QLabel("Vehicle Make"), 2, 0)
        grid.addWidget(self.m_vehicle_make_edit, 2, 1)
        grid.addWidget(QLabel("Model"), 2, 2)
        grid.addWidget(self.m_vehicle_model_edit, 2, 3)

        # Row 3: Lic. Plate #
        grid.addWidget(QLabel("Lic. Plate #"), 3, 0)
        grid.addWidget(self.m_plate_edit, 3, 1)

        # Row 4: Begin / End date
        grid.addWidget(QLabel("Begin Date"), 4, 0)
        grid.addWidget(self.m_begin_date_edit, 4, 1)
        grid.addWidget(QLabel("End Date"), 4, 2)
        grid.addWidget(self.m_end_date_edit, 4, 3)

        root.addWidget(header_box)

        # Table setup: column names correspond to database fields
        self.columns: List[str] = [
            "id",
            "date",
            "start_miles",
            "end_miles",
            "miles_driven",
            "start_location",
            "end_location",
            "purpose",
            "actions",
        ]

        # Labels shown in the header row of the table
        self.column_labels = {
            "id": "ID",
            "date": "Date",
            "start_miles": "Start",
            "end_miles": "End",
            "miles_driven": "Miles",
            "start_location": "Start Location",
            "end_location": "End Location",
            "purpose": "Purpose",
            "actions": "Actions",
        }

        # The table view widget
        self.table = BaseTableView(self)
        root.addWidget(self.table, stretch=1)

        # Base model and filter proxy
        self.base_model = MileageTableModel([], self.columns, self.column_labels)
        self.proxy = MileageFilterProxy(self.columns, self)
        self.proxy.setSourceModel(self.base_model)

        # Connect model to table
        self.table.setModel(self.proxy)

        # Hide database ID column
        id_col = self.columns.index("id")
        self.table.setColumnHidden(id_col, True)

        self.table.setSortingEnabled(True)

        # Action buttons (edit/delete)
        self.actions_delegate = ActionButtonsDelegate(
            parent=self.table,
            button_keys=["edit", "delete"],
            on_edit=self._on_edit_clicked,
            on_delete=self._on_delete_clicked,
        )

        actions_col = self.columns.index("actions")
        self.table.setItemDelegateForColumn(actions_col, self.actions_delegate)
        self._ensure_actions_width()

        # Ensure actions column is wide enough
        QTimer.singleShot(0, self._ensure_actions_width)
        self.base_model.modelReset.connect(self._ensure_actions_width)
        self.base_model.layoutChanged.connect(self._ensure_actions_width)

        # Double click > Edit a record
        self.table.doubleClicked.connect(self._on_double_click)

        # Load all mileage entries from DB
        self.load_items()

    # Load items into table
    def load_items(self):
        items = MileageEntry.get_all_for_user(self.user)
        self.base_model.set_items(items)
        self.table.resizeColumnsToContents()

        # Ensure actions column resizes correctly after loading
        QTimer.singleShot(0, self._ensure_actions_width)

    def refresh_table(self):
        self.load_items()

    # Helper: map rows/indexes to db items
    def _map_proxy_index_to_item(self, proxy_index: QModelIndex) -> Optional[MileageEntry]:
        if not proxy_index.isValid():
            return None
        source_index = self.proxy.mapToSource(proxy_index)
        return self.base_model.get_item(source_index.row())

    def _map_source_row_to_item(self, source_row: int) -> Optional[MileageEntry]:
        return self.base_model.get_item(source_row)

    # Delegate callbacks (edit/delete buttons)
    def _on_edit_clicked(self, source_row: int):
        item = self._map_source_row_to_item(source_row)
        if item:
            self.edit_item(item)

    def _on_delete_clicked(self, source_row: int):
        item = self._map_source_row_to_item(source_row)
        if item:
            self.delete_item(item)

    # Forms: Add/Edit mileage entry
    def open_add_form(self):
        form = MileageForm(self, on_save=self.refresh_table, item=None)
        form.exec()

    def edit_item(self, item: MileageEntry):
        form = MileageForm(self, on_save=self.refresh_table, item=item)
        form.exec()

    # Delete mileage entry
    def delete_item(self, item: MileageEntry):
        confirm = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete mileage entry for {item.date}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if confirm != QMessageBox.Yes:
            return

        MileageEntry.delete(item.id, self.user)
        self.load_items()

    # Double-click handler > edit entry
    def _on_double_click(self, index: QModelIndex):
        item = self._map_proxy_index_to_item(index)
        if item:
            self.edit_item(item)

    # Simple filtering dialog (substring filter)
    def open_filter_dialog(self):
        # Opens a dialog to apply a simple substring filter to a chosen column.
        dlg = QDialog(self)
        dlg.setWindowTitle("Filter Mileage")

        form = QFormLayout(dlg)

        col_combo = QComboBox(dlg)
        # Allow filtering on every column except the action buttons
        filterable_keys = [c for c in self.columns if c != "actions"]

        for key in filterable_keys:
            col_combo.addItem(self.column_labels.get(key, key), key)

        text_edit = QLineEdit(dlg)

        form.addRow("Column:", col_combo)
        form.addRow("Contains:", text_edit)

        # OK = apply, Cancel = do nothing, Reset = clear all filters
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel | QDialogButtonBox.Reset,
            parent=dlg,
        )
        form.addRow(buttons)

        # Reset filters button
        def on_reset():
            self.proxy.clear_all_filters()
            text_edit.clear()

        buttons.button(QDialogButtonBox.Reset).clicked.connect(on_reset)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)

        # When user pressed OK
        if dlg.exec() == QDialog.Accepted:
            key = col_combo.currentData()
            text = text_edit.text()
            self.proxy.set_text_filter(key, text)

    def export_to_excel(self):
    # Exports the mileage entries and header info into the Excel report template.

        # Ensure data is in entry
        if self.base_model.rowCount() == 0:
            QMessageBox.information(self, "No Data", "There are no mileage entries to export.")
            return

        # Ask user where to save the output file
        month_name = datetime.now().strftime("%B").lower()
        default_filename = f"{month_name}_mileage_report_{self.user}.xlsx"

        out_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Mileage Expense Report",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*.*)",
        )
        if not out_path:
            return

        # Path to users preformatted template
        template_path = os.path.join("assets", "Forms", "MileageExpenseFormBlank.xlsx")

        if not os.path.exists(template_path):
            QMessageBox.critical(
                self,
                "Template Missing",
                f"Could not find template file:\n{template_path}\n\n"
                "Update template_path in MileagePage.export_to_excel(); Ensure the template "
                "you've chosen is MileageExpenseFormBlank.xlsx and not MileageExpenseFormBlank.xls"
            )
            return

        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # Read header info from widgets
            emp_name = self.m_emp_name_edit.text().strip()
            emp_number = self.m_emp_number_edit.text().strip()
            phone = self.m_phone_edit.text().strip()
            mail_team = self.m_mail_team_edit.text().strip()

            vehicle_make = self.m_vehicle_make_edit.text().strip()
            vehicle_model = self.m_vehicle_model_edit.text().strip()
            plate = self.m_plate_edit.text().strip()

            begin_q = self.m_begin_date_edit.date()
            end_q = self.m_end_date_edit.date()

            begin_date = date(begin_q.year(), begin_q.month(), begin_q.day())
            end_date = date(end_q.year(), end_q.month(), end_q.day())

            # Write header cells
            ws["A3"] = emp_name
            ws["C3"] = emp_number
            ws["F3"] = phone
            ws["H3"] = mail_team

            ws["A6"] = vehicle_make
            ws["C6"] = vehicle_model
            ws["D6"] = plate
            ws["F6"] = begin_date.strftime("%m/%d/%Y") if begin_date else ""
            ws["G6"] = end_date.strftime("%m/%d/%Y") if end_date else ""

            # Collect entries in date range
            def to_date(val):
                if isinstance(val, date):
                    return val
                if isinstance(val, datetime):
                    return val.date()
                if isinstance(val, str) and val.strip():
                    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                        try:
                            return datetime.strptime(val, fmt).date()
                        except ValueError:
                            continue
                return None

            export_items = []
            for row in range(self.base_model.rowCount()):
                entry = self.base_model.get_item(row)
                d = to_date(getattr(entry, "date", None))
                if d and begin_date <= d <= end_date:
                    export_items.append((entry, d))

            if not export_items:
                QMessageBox.information(
                    self,
                    "No Items",
                    "No mileage entries fall within the selected date range.",
                )
                return

            # Write line items starting at row 9
            start_row = 9
            for idx, (entry, d) in enumerate(export_items):
                r = start_row + idx

                ws[f"A{r}"] = d.strftime("%m/%d/%Y") if d else ""
                ws[f"B{r}"] = getattr(entry, "start_miles", "") or 0
                ws[f"C{r}"] = getattr(entry, "end_miles", "") or 0

                # Let the template compute # Miles and Mileage Exp (columns D/E)

                # From / To / Purpose
                ws[f"H{r}"] = getattr(entry, "start_location", "") or ""
                ws[f"I{r}"] = getattr(entry, "end_location", "") or ""
                ws[f"M{r}"] = getattr(entry, "purpose", "") or ""

            # Save the filled-out workbook
            wb.save(out_path)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Mileage expense report exported to:\n{out_path}",
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"An error occurred while exporting:\n{e}",
            )

    # Clear all entries
    def clear_all_entries(self):
        confirm = QMessageBox.question(
            self,
            "Confirm Deletion",
            "This will permanently delete all stored mileage entries.\n\n"
            "Are you sure you want to continue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if confirm != QMessageBox.Yes:
            return

        try:
            clear_all_mileage_entries(self.user)
            self.load_items()

            QMessageBox.information(
                self,
                "Mileage Tracker Reset",
                "All mileage entries have been cleared.",
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to clear entries:\n{e}")

    # Action columns width
    def _actions_logical_index(self) -> int:
        try:
            return self.columns.index("actions")
        except ValueError:
            return -1

    def _ensure_actions_width(self):
        # Ensures the actions column width matches the delegate's requirements.
        li = self._actions_logical_index()
        if li < 0:
            return
        header = self.table.horizontalHeader()
        delegate = self.table.itemDelegateForColumn(li)
        if hasattr(delegate, "required_width"):
            width = delegate.required_width()
        else:
            width = 220
        header.setSectionResizeMode(li, QHeaderView.ResizeMode.Interactive)
        if header.sectionSize(li) < width:
            header.resizeSection(li, width)

    def show_help_dialog(self):
        dlg = MileageHelpDialog(self)
        dlg.exec()